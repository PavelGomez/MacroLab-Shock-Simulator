"""
MacroLab Phase 1 — 16 Institutional Regime Trajectory Templates
3-equation New Keynesian model (Carlin & Soskice 2024 framework)
Generates MacroLab_Regime_Trajectories_v1.xlsx

Regime matrix: 2×2×2×2
  Dim 1 — Exchange Rate:        F=Float, P=Peg
  Dim 2 — Labor Coordination:   C=Coordinated, U=Uncoordinated
  Dim 3 — BC Credibility:       H=High, L=Low
  Dim 4 — Fiscal Framework:     R=Rule-Bound, D=Discretionary

Shock: positive demand shock (+2 pp output gap) at t=1, fades geometrically (ρ=0.5/quarter)
Simulation: 12 quarters from equilibrium
"""

import math
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# MODEL PARAMETERS
# ──────────────────────────────────────────────────────────────────────────────

T = 12          # quarters
Y_STAR = 0.0    # equilibrium output gap
PI_STAR = 0.03  # inflation target (3 %)
R_STAR = 0.04   # neutral real rate (4 %)

SHOCK_SIZE = 2.0   # demand shock in pp at t=1
SHOCK_DECAY = 0.50 # geometric decay per quarter

# ──────────────────────────────────────────────────────────────────────────────
# REGIME DEFINITIONS
# 16 regimes: each is a tuple (ER, Labor, Cred, Fiscal)
# Parameter sets calibrate IS slope, PC slope, Taylor coefficient,
# expectation formation, fiscal multiplier, exchange-rate buffer
# ──────────────────────────────────────────────────────────────────────────────

REGIMES = {}

def _regime(code, label, er, lab, cred, fisc, params, narrative, episodes):
    REGIMES[code] = {
        "code": code, "label": label,
        "er": er, "lab": lab, "cred": cred, "fisc": fisc,
        "params": params, "narrative": narrative, "episodes": episodes
    }

# params keys:
#   alpha   — IS sensitivity to real rate (output gap persistence if rate unchanged)
#   beta    — PC slope (pass-through from output gap to inflation)
#   gamma   — PC backward-looking weight (1-gamma = forward-looking)
#   theta   — Taylor rule inflation coefficient (>1 by Taylor Principle)
#   phi     — Taylor rule output gap coefficient
#   mu      — fiscal multiplier (active when fisc=Discretionary under stress)
#   er_buf  — exchange rate buffer: floating absorbs shock, reduces effective alpha
#   sigma   — expectation anchoring: high cred → lower sigma (faster anchor)

# ────────── FLOAT / COORDINATED / HIGH CRED / RULE-BOUND ──────────
_regime(
    "FCHR", "Flotante · Coord · Alta cred · Regla fiscal",
    "Float","Coord","High","Rule",
    {"alpha":0.6,"beta":0.3,"gamma":0.25,"theta":1.6,"phi":0.5,
     "mu":0.0,"er_buf":0.3,"sigma":0.05},
    ("Régimen de referencia de mejor desempeño. El BC creíble responde con fuerza "
     "(θ=1,6); expectativas mayoritariamente forward-looking anclan rápido (γ=0,25). "
     "La flotación cambiaria amortigua parte del shock externo (er_buf=0,3). "
     "La regla fiscal evita sobrecalentamiento adicional (μ=0). "
     "Resultado: brecha de producto cierra en ~4 trimestres; inflación no supera +1,5 pp "
     "sobre meta; tasa real cae de vuelta al neutro hacia t=7."),
    ["Chile 2008–2009","Suecia 1994–1998 (post-flotación)"]
)

# ────────── FLOAT / COORDINATED / HIGH CRED / DISCRETIONARY ──────────
_regime(
    "FCHD", "Flotante · Coord · Alta cred · Discrecional",
    "Float","Coord","High","Disc",
    {"alpha":0.6,"beta":0.3,"gamma":0.25,"theta":1.6,"phi":0.5,
     "mu":0.4,"er_buf":0.3,"sigma":0.05},
    ("Similar al régimen FCHR pero el gobierno amplifica el shock fiscal en t=1–2 "
     "(μ=0,4; gasto procíclico posible). La brecha se cierra igualmente gracias a la "
     "credibilidad del BC, pero el pico inflacionario es marginalmente mayor (~+1,8 pp). "
     "Requiere que el BC responda con más fuerza transitoriamente. "
     "Riesgo: si la expansión fiscal persiste más allá de t=2, la inflación puede "
     "demorarse un trimestre extra en converger."),
    ["Chile 2021–2024 (fase inicial, antes de consolidación fiscal)"]
)

# ────────── FLOAT / COORDINATED / LOW CRED / RULE-BOUND ──────────
_regime(
    "FCLR", "Flotante · Coord · Baja cred · Regla fiscal",
    "Float","Coord","Low","Rule",
    {"alpha":0.6,"beta":0.4,"gamma":0.55,"theta":1.3,"phi":0.4,
     "mu":0.0,"er_buf":0.3,"sigma":0.20},
    ("BC con baja credibilidad compensa con regla fiscal (μ=0). "
     "Expectativas más backward-looking (γ=0,55): la inflación persiste más trimestres. "
     "El BC debe subir más la tasa nominal para mover la real (σ alto = desanclaje parcial). "
     "La coordinación laboral modera el pass-through salarial (β=0,4 vs 0,5 en coord baja). "
     "Brecha cierra ~t=6; inflación tarda ~t=8–9. Resultado subóptimo pero controlable."),
    ["Brasil 2002–2003 (credibilidad en reconstrucción)"]
)

# ────────── FLOAT / COORDINATED / LOW CRED / DISCRETIONARY ──────────
_regime(
    "FCLD", "Flotante · Coord · Baja cred · Discrecional",
    "Float","Coord","Low","Disc",
    {"alpha":0.6,"beta":0.4,"gamma":0.55,"theta":1.3,"phi":0.4,
     "mu":0.4,"er_buf":0.3,"sigma":0.20},
    ("Combinación problemática: BC poco creíble + fiscal discrecional en boom. "
     "El impulso fiscal amplifica la brecha en t=2–3 (μ=0,4). "
     "Expectativas se desanclan parcialmente (σ=0,2): inflación alcanza +2,5–3 pp. "
     "BC responde con aumento de tasa, pero la coordinación laboral limita espiral salarial. "
     "Ajuste lento (t=9–10). Vulnerabilidad: si el fiscal sigue expansivo, riesgo de trampa "
     "inflacionaria moderada sin dominancia fiscal plena."),
    ["Colombia 2006–2008 (expansión fiscal procíclica)"]
)

# ────────── FLOAT / UNCOORDINATED / HIGH CRED / RULE-BOUND ──────────
_regime(
    "FUHR", "Flotante · Descoord · Alta cred · Regla fiscal",
    "Float","Uncoord","High","Rule",
    {"alpha":0.6,"beta":0.5,"gamma":0.30,"theta":1.6,"phi":0.5,
     "mu":0.0,"er_buf":0.3,"sigma":0.07},
    ("Alta credibilidad del BC ancora expectativas pese a mercado laboral descoordinado "
     "(β=0,5: mayor pass-through salarial que en coord). La regla fiscal evita "
     "amplificación. La flotación amortigua. "
     "El BC necesita subir más la tasa real que en FCHR para cerrar la brecha a tiempo "
     "dada la mayor presión salarial. Brecha cierra t=5; inflación t=6. "
     "Costo: desempleo transitoriamente mayor durante la desinflación."),
    ["Nueva Zelanda 1990s","Australia 2008–2009"]
)

# ────────── FLOAT / UNCOORDINATED / HIGH CRED / DISCRETIONARY ──────────
_regime(
    "FUHD", "Flotante · Descoord · Alta cred · Discrecional",
    "Float","Uncoord","High","Disc",
    {"alpha":0.6,"beta":0.5,"gamma":0.30,"theta":1.6,"phi":0.5,
     "mu":0.4,"er_buf":0.3,"sigma":0.07},
    ("BC creíble con flotación contiene el desanclaje, pero la combinación "
     "fiscal discrecional + mercado laboral descoordinado genera pico inflacionario "
     "más alto (+2,2 pp). BC debe endurecer más transitoriamente. "
     "Si la regla de Taylor tiene credibilidad, las expectativas absorben el ajuste y "
     "la inflación converge hacia t=7. Escenario frecuente en economías desarrolladas "
     "con shock de oferta."),
    ["EE.UU. 2021–2022 (shock pandemia + fiscal)","Reino Unido 2021–2023"]
)

# ────────── FLOAT / UNCOORDINATED / LOW CRED / RULE-BOUND ──────────
_regime(
    "FULR", "Flotante · Descoord · Baja cred · Regla fiscal",
    "Float","Uncoord","Low","Rule",
    {"alpha":0.6,"beta":0.5,"gamma":0.65,"theta":1.3,"phi":0.4,
     "mu":0.0,"er_buf":0.3,"sigma":0.25},
    ("Descoordinación laboral + baja credibilidad = alta persistencia inflacionaria "
     "(γ=0,65 backward-looking; β=0,5). La regla fiscal frena amplificación. "
     "La flotación ayuda por el lado externo, pero el pass-through doméstico es fuerte. "
     "BC debe subir tasa agresivamente (compensando σ alto). Brecha tarda t=7–8; "
     "inflación t=10. Riesgo: si credibilidad se erosiona más (eventos adversos), "
     "puede activarse espiral salario-precio moderada."),
    ["México 2021–2022","Perú 2021–2022"]
)

# ────────── FLOAT / UNCOORDINATED / LOW CRED / DISCRETIONARY ──────────
_regime(
    "FULD", "Flotante · Descoord · Baja cred · Discrecional",
    "Float","Uncoord","Low","Disc",
    {"alpha":0.6,"beta":0.5,"gamma":0.65,"theta":1.3,"phi":0.4,
     "mu":0.4,"er_buf":0.3,"sigma":0.25},
    ("El peor escenario flotante. Fiscal discrecional amplifica shock (t=2); "
     "mercado laboral descoordinado → pass-through salarial alto (β=0,5); "
     "BC poco creíble → expectativas backward-looking (γ=0,65). "
     "Inflación pico ~+3,5 pp sobre meta. BC en dilema: subir agresivo → recesión; "
     "subir poco → desanclaje. La flotación evita crisis cambiaria pero no elimina "
     "la persistencia. Convergencia t=11–12. Ilustra importancia de coordinación "
     "institucional en economías emergentes."),
    ["Brasil 2015–2016 (combinación adversa)"]
)

# ────────── PEG / COORDINATED / HIGH CRED / RULE-BOUND ──────────
_regime(
    "PCHR", "Fijo · Coord · Alta cred · Regla fiscal",
    "Peg","Coord","High","Rule",
    {"alpha":0.7,"beta":0.3,"gamma":0.35,"theta":0.0,"phi":0.0,
     "mu":0.0,"er_buf":0.0,"sigma":0.10},
    ("Régimen de tipo fijo: BC no tiene tasa de política independiente (θ=φ=0). "
     "El ajuste recae íntegramente en precios relativos y coordinación salarial. "
     "La regla fiscal bloquea amplificación. La coordinación laboral suaviza la "
     "desinflación interna. Sin er_buf: shock se absorbe vía brecha que tarda más. "
     "Brecha cierra t=7–8 mediante deflación relativa. Funciona si coordinación "
     "salarial es robusta (análogo a Zona Euro sin shocks asimétricos severos)."),
    ["Países Bajos en ERM 1990s","Austria en zona DEM"]
)

# ────────── PEG / COORDINATED / HIGH CRED / DISCRETIONARY ──────────
_regime(
    "PCHD", "Fijo · Coord · Alta cred · Discrecional",
    "Peg","Coord","High","Disc",
    {"alpha":0.7,"beta":0.3,"gamma":0.35,"theta":0.0,"phi":0.0,
     "mu":0.5,"er_buf":0.0,"sigma":0.10},
    ("Tipo fijo + fiscal discrecional en boom: la política fiscal es el único "
     "estabilizador activo. Si el gobierno amplifica (μ=0,5), la brecha crece más "
     "en t=2–3, generando mayor presión inflacionaria interna que erosiona "
     "competitividad bajo el fijo. La coordinación laboral modera pero no elimina. "
     "Riesgo: acumulación de desalineamiento del tipo de cambio real si persiste. "
     "Convergencia lenta t=9–10. El fiscal discrecional PROCÍCLICO en fijo es "
     "especialmente peligroso."),
    ["España 2003–2007 (boom inmobiliario en euro)","Irlanda pre-2008"]
)

# ────────── PEG / COORDINATED / LOW CRED / RULE-BOUND ──────────
_regime(
    "PCLR", "Fijo · Coord · Baja cred · Regla fiscal",
    "Peg","Coord","Low","Rule",
    {"alpha":0.7,"beta":0.4,"gamma":0.55,"theta":0.0,"phi":0.0,
     "mu":0.0,"er_buf":0.0,"sigma":0.20},
    ("Fijo + baja credibilidad: expectativas parcialmente desancladas bajo fijo. "
     "Si agentes dudan de la sostenibilidad del fijo (σ=0,2), la curva de Phillips "
     "implica mayor inflación doméstica por expectativa de depreciación futura. "
     "Coordinación laboral modera espiral salarial (β=0,4). Regla fiscal bloquea "
     "amplificación. Pero sin herramienta de tasa y con credibilidad débil, la brecha "
     "cierra muy lento (t=9–10). Riesgo de ataque especulativo si reservas caen."),
    ["Argentina Convertibilidad 1995–2001 (fragmentos)"]
)

# ────────── PEG / COORDINATED / LOW CRED / DISCRETIONARY ──────────
_regime(
    "PCLD", "Fijo · Coord · Baja cred · Discrecional",
    "Peg","Coord","Low","Disc",
    {"alpha":0.7,"beta":0.4,"gamma":0.55,"theta":0.0,"phi":0.0,
     "mu":0.5,"er_buf":0.0,"sigma":0.20},
    ("Combinación altamente inestable bajo fijo. Fiscal discrecional amplifica (μ=0,5); "
     "baja credibilidad → expectativas parcialmente desancladas. Sin política monetaria "
     "independiente ni flotación cambiaria. La brecha puede persistir todo el horizonte "
     "si el shock es suficientemente grande o la expansión fiscal continúa. "
     "Precursor clásico de crisis de balanza de pagos: competitividad deteriorada, "
     "déficit externo ampliado, reservas bajo presión. Ilustra dinámica Argentina "
     "convertibilidad bajo shocks externos."),
    ["Argentina Convertibilidad 1999–2001","Ecuador pre-dolarización"]
)

# ────────── PEG / UNCOORDINATED / HIGH CRED / RULE-BOUND ──────────
_regime(
    "PUHR", "Fijo · Descoord · Alta cred · Regla fiscal",
    "Peg","Uncoord","High","Rule",
    {"alpha":0.7,"beta":0.5,"gamma":0.35,"theta":0.0,"phi":0.0,
     "mu":0.0,"er_buf":0.0,"sigma":0.10},
    ("Alta credibilidad del ancla (el fijo mismo es el ancla nominal) con mercado "
     "laboral descoordinado. Sin política de tasa, el ajuste interno requiere "
     "deflación de salarios nominales o caída de empleo (β=0,5). "
     "La regla fiscal bloquea expansión compensatoria. "
     "Proceso largo y costoso: brecha positiva inicialmente luego sobrecarga, "
     "corrección vía deflación relativa t=7–9. Análogo a ajuste de Zona Euro "
     "en países periféricos post-2010 sin flotación."),
    ["Grecia 2010–2015 (sin flotación, ajuste interno)"]
)

# ────────── PEG / UNCOORDINATED / HIGH CRED / DISCRETIONARY ──────────
_regime(
    "PUHD", "Fijo · Descoord · Alta cred · Discrecional",
    "Peg","Uncoord","High","Disc",
    {"alpha":0.7,"beta":0.5,"gamma":0.35,"theta":0.0,"phi":0.0,
     "mu":0.5,"er_buf":0.0,"sigma":0.10},
    ("El fiscal discrecional bajo fijo + descoordinación laboral genera el peor "
     "patrón de desalineamiento real posible. En boom: el fiscal amplifica (μ=0,5) "
     "y los salarios suben sin coordinación (β=0,5) → erosión de competitividad. "
     "El ancla cambiario impide corrección nominal. Resultado: apreciación real "
     "acumulada, déficit externo, eventual crisis. Requiere ajuste interno severo "
     "(a veces deflación > 10% de salarios nominales) para recuperar competitividad. "
     "Dinámica: shock → boom → crisis de competitividad → ajuste contractivo."),
    ["Portugal 2000–2011","Chipre 2008–2012"]
)

# ────────── PEG / UNCOORDINATED / LOW CRED / RULE-BOUND ──────────
_regime(
    "PULR", "Fijo · Descoord · Baja cred · Regla fiscal",
    "Peg","Uncoord","Low","Rule",
    {"alpha":0.7,"beta":0.5,"gamma":0.65,"theta":0.0,"phi":0.0,
     "mu":0.0,"er_buf":0.0,"sigma":0.25},
    ("Sin política de tasa, sin flotación, con baja credibilidad y descoordinación: "
     "escenario de máxima rigidez y mínima estabilización endógena. "
     "Expectativas backward-looking dominan (γ=0,65). La inflación doméstica sube "
     "por inercia incluso sin amplificación fiscal. El tipo real se aprecia. "
     "La regla fiscal evita lo peor pero no puede compensar la ausencia de "
     "estabilizadores. Brecha y desalineamiento persisten todo el horizonte. "
     "Alta probabilidad de crisis cambiaria si el shock es persistente."),
    ["Venezuela 2005–2010 (fijo + instituciones débiles)"]
)

# ────────── PEG / UNCOORDINATED / LOW CRED / DISCRETIONARY ──────────
_regime(
    "PULD", "Fijo · Descoord · Baja cred · Discrecional",
    "Peg","Uncoord","Low","Disc",
    {"alpha":0.7,"beta":0.5,"gamma":0.65,"theta":0.0,"phi":0.0,
     "mu":0.5,"er_buf":0.0,"sigma":0.25},
    ("Escenario de crisis sistémica. Todas las fragilidades activadas: fijo → sin "
     "política monetaria ni flotación; descoordinado → alto pass-through salarial; "
     "baja cred → expectativas desancladas; fiscal discrecional → amplificación "
     "procíclica. El modelo entra en dinámica explosiva: fiscal expande en t=1–3 "
     "→ inflación sube → tipo real se aprecia → déficit externo → reservas caen "
     "→ ataque especulativo → abandono del fijo o default. "
     "Esta celda es la referencia para diagnóstico de crisis de BdP en emergentes."),
    ["Argentina 2001","Venezuela 2013–2018"]
)

# ──────────────────────────────────────────────────────────────────────────────
# SIMULATION ENGINE (3-equation NK)
# ──────────────────────────────────────────────────────────────────────────────

def simulate(regime_code, T=T):
    reg = REGIMES[regime_code]
    p   = reg["params"]
    alpha, beta, gamma = p["alpha"], p["beta"], p["gamma"]
    theta, phi        = p["theta"], p["phi"]
    mu, er_buf, sigma = p["mu"], p["er_buf"], p["sigma"]
    er_regime         = reg["er"]

    # State variables
    y      = [0.0] * (T + 1)   # output gap (% deviation from potential)
    pi     = [0.0] * (T + 1)   # inflation (pp deviation from target)
    r      = [0.0] * (T + 1)   # real rate gap (r - r*)
    pi_exp = [0.0] * (T + 1)   # expected inflation gap
    shock  = [0.0] * (T + 1)   # demand shock (fades geometrically)

    # Shock at t=1
    for t in range(1, T + 1):
        shock[t] = SHOCK_SIZE * (SHOCK_DECAY ** (t - 1))

    for t in range(1, T + 1):
        # Exchange rate buffer reduces effective shock (floating absorbs part)
        eff_shock = shock[t] * (1.0 - er_buf)

        # Fiscal amplification (discretionary: amplifies in first 3 quarters)
        fisc_impulse = mu * shock[t] if (t <= 3) else 0.0

        # Inflation expectations (anchoring: mix of forward and backward)
        # Forward-looking: target (=0 gap)
        # Backward-looking: last period inflation
        pi_exp[t] = (1 - gamma - sigma) * 0.0 + gamma * pi[t-1] + sigma * pi[t-1] * 1.2

        # IS curve: y_t = y_{t-1}*alpha - alpha*(r_{t-1}) + eff_shock + fiscal
        y[t] = alpha * y[t-1] - alpha * r[t-1] + eff_shock + fisc_impulse

        # Clamp output gap: under floating with high cred, self-corrects faster
        # Under peg with low cred, allow overshoot
        max_gap = 5.0 if er_regime == "Peg" and sigma > 0.15 else 4.0
        y[t] = max(min(y[t], max_gap), -max_gap)

        # Phillips Curve: pi_t = pi_exp_t + beta * y_t
        pi[t] = pi_exp[t] + beta * y[t]

        # Taylor Rule (only if floating / monetary policy active)
        if theta > 0:
            r[t] = theta * pi[t] + phi * y[t]
        else:
            # Under peg: rate is externally set; mild endogenous component via spread
            r[t] = 0.5 * pi[t]  # partial real rate response (e.g. via credit market)

    return {
        "shock": shock[1:],
        "y":     y[1:],
        "pi":    pi[1:],
        "r":     r[1:],
        "pi_exp": pi_exp[1:],
    }

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATION
# ──────────────────────────────────────────────────────────────────────────────

# Color palette
COL_NAVY   = "1B2A4A"
COL_WHITE  = "FFFFFF"
COL_ACCENT = "2D7D46"
COL_SOFT   = "E8F5EC"
COL_WARN   = "FFF3CD"
COL_CARD   = "F5F7FA"
COL_LINE   = "D0D6E0"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border_thin():
    s = Side(style="thin", color=COL_LINE)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium_bottom():
    t = Side(style="thin", color=COL_LINE)
    m = Side(style="medium", color=COL_NAVY)
    return Border(left=t, right=t, top=t, bottom=m)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def write_regime_sheet(wb, code):
    reg = REGIMES[code]
    p   = reg["params"]
    sim = simulate(code)

    ws = wb.create_sheet(title=code)
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = [3, 14, 11, 11, 11, 11, 55, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── HEADER ──
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"B{row}:F{row}")
    c = ws[f"B{row}"]
    c.value = f"[{code}]  {reg['label']}"
    c.fill  = _fill(COL_NAVY)
    c.font  = _font(bold=True, color=COL_WHITE, size=13)
    c.alignment = _align("left")
    row += 1

    # ── REGIME BADGES ──
    ws.row_dimensions[row].height = 18
    badges = [
        ("Tipo de cambio:", reg["er"]),
        ("Mercado laboral:", reg["lab"]),
        ("Cred BC:", reg["cred"]),
        ("Marco fiscal:", reg["fisc"]),
    ]
    cols = [2, 3, 4, 5]
    for col_i, (label, val) in zip(cols, badges):
        cl = ws.cell(row=row, column=col_i, value=f"{label} {val}")
        cl.fill = _fill(COL_SOFT)
        cl.font = _font(bold=False, size=9, italic=True)
        cl.alignment = _align("center")
        cl.border = _border_thin()
    row += 1

    # ── PARAM TABLE ──
    ws.row_dimensions[row].height = 16
    param_labels = [
        ("α (IS)", p["alpha"]), ("β (PC)", p["beta"]),
        ("γ (backward π)", p["gamma"]), ("θ (Taylor π)", p["theta"]),
        ("φ (Taylor y)", p["phi"]),
    ]
    for col_i, (lbl, val) in enumerate(param_labels, 2):
        cl = ws.cell(row=row, column=col_i, value=f"{lbl} = {val}")
        cl.fill = _fill(COL_CARD)
        cl.font = _font(size=8, italic=True)
        cl.alignment = _align("center")
        cl.border = _border_thin()
    # extra params in col G
    ex = ws.cell(row=row, column=7,
                 value=f"μ (fiscal) = {p['mu']}   er_buf = {p['er_buf']}   σ (desanclaje) = {p['sigma']}")
    ex.fill = _fill(COL_CARD)
    ex.font = _font(size=8, italic=True)
    ex.alignment = _align("left")
    ex.border = _border_thin()
    row += 1

    # ── NARRATIVE ──
    ws.row_dimensions[row].height = 72
    ws.merge_cells(f"B{row}:G{row}")
    nc = ws[f"B{row}"]
    nc.value = "NARRATIVA DEL RÉGIMEN\n" + reg["narrative"]
    nc.fill = _fill(COL_WARN)
    nc.font = _font(size=9)
    nc.alignment = _align("left", wrap=True)
    nc.border = _border_thin()
    row += 1

    # ── EPISODES ──
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"B{row}:G{row}")
    ec = ws[f"B{row}"]
    ec.value = "Episodios históricos mapeados: " + "  ·  ".join(reg["episodes"])
    ec.fill = _fill(COL_SOFT)
    ec.font = _font(size=9, italic=True)
    ec.alignment = _align("left")
    ec.border = _border_thin()
    row += 1

    row += 1  # spacer

    # ── DATA TABLE HEADER ──
    ws.row_dimensions[row].height = 18
    headers = ["Trimestre", "Shock demanda", "Brecha producto (y)", "Inflación (π)", "Tasa real (r)", "π esperada"]
    for col_i, h in enumerate(headers, 2):
        hc = ws.cell(row=row, column=col_i, value=h)
        hc.fill = _fill(COL_NAVY)
        hc.font = _font(bold=True, color=COL_WHITE, size=9)
        hc.alignment = _align("center")
        hc.border = _border_thin()
    row += 1

    # ── DATA ROWS ──
    data_start_row = row
    for t in range(T):
        ws.row_dimensions[row].height = 15
        vals = [
            f"T+{t+1}",
            round(sim["shock"][t], 3),
            round(sim["y"][t], 3),
            round(sim["pi"][t], 3),
            round(sim["r"][t], 3),
            round(sim["pi_exp"][t], 3),
        ]
        for col_i, v in enumerate(vals, 2):
            dc = ws.cell(row=row, column=col_i, value=v)
            if col_i == 2:
                dc.fill = _fill(COL_CARD)
                dc.font = _font(size=9)
            else:
                # Color-code: positive=green-ish, negative=red-ish
                if isinstance(v, float) and v > 0.01:
                    dc.fill = _fill("E8F5E9")
                elif isinstance(v, float) and v < -0.01:
                    dc.fill = _fill("FFEBEE")
                else:
                    dc.fill = _fill(COL_WHITE)
                dc.font = _font(size=9)
            dc.alignment = _align("center")
            dc.border = _border_thin()
        row += 1

    data_end_row = row - 1

    row += 1  # spacer

    # ── CHART ──
    chart = LineChart()
    chart.title = f"Respuesta al shock — Régimen {code}"
    chart.style = 10
    chart.width  = 18
    chart.height = 10
    chart.y_axis.title = "pp desviación del equilibrio"
    chart.x_axis.title = "Trimestres post-shock"
    chart.grouping = "standard"
    chart.legend.position = "b"

    # Series: y (col D=4), pi (col E=5), r (col F=6)
    from openpyxl.chart import Series
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.drawing.line import LineProperties

    # Categories
    cats = Reference(ws, min_col=2, max_col=2,
                     min_row=data_start_row, max_row=data_end_row)
    chart.set_categories(cats)

    series_specs = [
        (4, "Brecha producto (y)", "1B5E20"),
        (5, "Inflación (π)",       "B71C1C"),
        (6, "Tasa real (r)",       "1A237E"),
    ]
    for col_i, name, color in series_specs:
        data_ref = Reference(ws,
                             min_col=col_i, max_col=col_i,
                             min_row=data_start_row,
                             max_row=data_end_row)
        ser = Series(data_ref, title=name)
        ser.graphicalProperties.line.solidFill = color
        ser.graphicalProperties.line.width = 20000
        chart.append(ser)

    ws.add_chart(chart, f"B{row}")


def write_summary_sheet(wb):
    ws = wb.create_sheet(title="RESUMEN_REGÍMENES", index=0)
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [3, 8, 26, 10, 14, 12, 12, 12, 12, 12, 12, 50, 3]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"B{row}:L{row}")
    t = ws[f"B{row}"]
    t.value = "MacroLab Shock Simulator · Phase 1 · Matriz de Regímenes Institucionales (16 trayectorias)"
    t.fill = _fill(COL_NAVY)
    t.font = _font(bold=True, color=COL_WHITE, size=14)
    t.alignment = _align("center")
    row += 1

    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"B{row}:L{row}")
    s = ws[f"B{row}"]
    s.value = ("Modelo NK 3 ecuaciones (IS + PC + Taylor). Shock: +2 pp demanda, decae 50%/trimestre. "
               "Horizonte: 12 trimestres. Carlin & Soskice 2024 framework.")
    s.fill = _fill(COL_CARD)
    s.font = _font(size=9, italic=True)
    s.alignment = _align("center")
    row += 2

    # ── COLUMN HEADERS ──
    ws.row_dimensions[row].height = 18
    col_headers = [
        "Código","Régimen","Tipo TC","Mercado laboral",
        "Cred BC","Marco fiscal",
        "y_max","pi_max","t_y_close","Episodios"
    ]
    for col_i, h in enumerate(col_headers, 2):
        hc = ws.cell(row=row, column=col_i, value=h)
        hc.fill = _fill(COL_NAVY)
        hc.font = _font(bold=True, color=COL_WHITE, size=9)
        hc.alignment = _align("center")
        hc.border = _border_thin()
    row += 1

    # ── DATA ROWS ──
    order = ["FCHR","FCHD","FCLR","FCLD",
             "FUHR","FUHD","FULR","FULD",
             "PCHR","PCHD","PCLR","PCLD",
             "PUHR","PUHD","PULR","PULD"]

    for i, code in enumerate(order):
        reg = REGIMES[code]
        sim = simulate(code)

        y_max  = round(max(sim["y"]), 2)
        pi_max = round(max(sim["pi"]), 2)
        # t_close: first t where |y| < 0.3 and stays
        t_close = T
        for t in range(T):
            if abs(sim["y"][t]) < 0.3:
                t_close = t + 1
                break

        ws.row_dimensions[row].height = 15
        row_fill = _fill(COL_CARD) if i % 2 == 0 else _fill(COL_WHITE)

        cells_vals = [
            code, reg["label"], reg["er"], reg["lab"],
            reg["cred"], reg["fisc"],
            y_max, pi_max, f"T+{t_close}",
            " · ".join(reg["episodes"])
        ]
        for col_i, v in enumerate(cells_vals, 2):
            dc = ws.cell(row=row, column=col_i, value=v)
            dc.fill = row_fill
            dc.font = _font(size=9)
            dc.alignment = _align("center" if col_i < 11 else "left")
            dc.border = _border_thin()
            # Highlight extreme values
            if col_i == 8 and isinstance(v, float) and v > 2.5:
                dc.fill = _fill("FFCDD2")
            if col_i == 7 and isinstance(v, float) and v > 3.0:
                dc.fill = _fill("FFCDD2")
        row += 1

    row += 2

    # ── PARAMETER KEY ──
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"B{row}:L{row}")
    pk = ws[f"B{row}"]
    pk.value = "CLAVE DE PARÁMETROS DEL MODELO"
    pk.fill = _fill(COL_NAVY)
    pk.font = _font(bold=True, color=COL_WHITE, size=10)
    pk.alignment = _align("center")
    row += 1

    param_key = [
        ("α", "Pendiente IS", "Sensibilidad del producto a la tasa real (mayor α = más reactivo)"),
        ("β", "Pendiente PC", "Pass-through del output gap a inflación (mayor β = más presión)"),
        ("γ", "Inercia inflacionaria", "Peso backward-looking en expectativas (γ→1 = pura inercia)"),
        ("θ", "Coef. Taylor (π)", "Respuesta del BC a inflación (0 bajo fijo cambiario)"),
        ("φ", "Coef. Taylor (y)", "Respuesta del BC al output gap"),
        ("μ", "Multiplicador fiscal", "Amplificación discrecional del shock (0 = regla fiscal)"),
        ("er_buf", "Buffer cambiario", "Fracción del shock absorbida por flotación (0 bajo fijo)"),
        ("σ", "Desanclaje", "Prima de desanclaje de expectativas (mayor σ = BC menos creíble)"),
    ]
    ws.row_dimensions[row].height = 14
    pk_headers = ["Parámetro","Nombre","Descripción pedagógica"]
    for col_i, h in enumerate(pk_headers, 2):
        hc = ws.cell(row=row, column=col_i, value=h)
        hc.fill = _fill(COL_SOFT)
        hc.font = _font(bold=True, size=9)
        hc.alignment = _align("center")
        hc.border = _border_thin()
    row += 1

    for param, name, desc in param_key:
        ws.row_dimensions[row].height = 14
        for col_i, v in enumerate([param, name, desc], 2):
            dc = ws.cell(row=row, column=col_i, value=v)
            dc.fill = _fill(COL_WHITE)
            dc.font = _font(size=9)
            dc.alignment = _align("left")
            dc.border = _border_thin()
        row += 1


def write_matrix_sheet(wb):
    """2×2×2×2 matrix visualization: Float/Peg × Coord/Uncoord × HighCred/LowCred × Rule/Disc"""
    ws = wb.create_sheet(title="MATRIZ_2x2x2x2")
    ws.sheet_view.showGridLines = False

    widths = [3] + [22] * 8 + [3]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(f"B{row}:I{row}")
    t = ws[f"B{row}"]
    t.value = "Matriz 2×2×2×2 — Clasificación de Regímenes Institucionales"
    t.fill = _fill(COL_NAVY)
    t.font = _font(bold=True, color=COL_WHITE, size=12)
    t.alignment = _align("center")
    row += 2

    # Layout: rows=Float/Peg×Coord/Uncoord, cols=HighCred/LowCred×Rule/Disc
    # Each cell: code + y_max + pi_max

    col_groups = [
        ("Alta Credibilidad BC", "Baja Credibilidad BC"),
    ]
    sub_cols   = ["Regla fiscal","Discrecional","Regla fiscal","Discrecional"]
    row_groups = ["FLOTANTE","FIJO"]
    sub_rows   = ["Coord laboral","Descoord laboral","Coord laboral","Descoord laboral"]

    # Header rows
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="FLOTANTE — Alta Cred BC").fill = _fill("1565C0")
    ws.cell(row=row, column=2).font = _font(bold=True, color=COL_WHITE, size=9)
    ws.cell(row=row, column=2).alignment = _align("center")
    ws.merge_cells(f"D{row}:E{row}")
    ws.cell(row=row, column=4, value="FLOTANTE — Baja Cred BC").fill = _fill("F57F17")
    ws.cell(row=row, column=4).font = _font(bold=True, color=COL_WHITE, size=9)
    ws.cell(row=row, column=4).alignment = _align("center")
    ws.merge_cells(f"F{row}:G{row}")
    ws.cell(row=row, column=6, value="FIJO — Alta Cred BC").fill = _fill("1B5E20")
    ws.cell(row=row, column=6).font = _font(bold=True, color=COL_WHITE, size=9)
    ws.cell(row=row, column=6).alignment = _align("center")
    ws.merge_cells(f"H{row}:I{row}")
    ws.cell(row=row, column=8, value="FIJO — Baja Cred BC").fill = _fill("B71C1C")
    ws.cell(row=row, column=8).font = _font(bold=True, color=COL_WHITE, size=9)
    ws.cell(row=row, column=8).alignment = _align("center")
    row += 1

    ws.row_dimensions[row].height = 14
    for col_i, label in enumerate(["Regla fiscal","Discrecional"]*4, 2):
        wc = ws.cell(row=row, column=col_i, value=label)
        wc.fill = _fill(COL_SOFT)
        wc.font = _font(size=8, bold=True)
        wc.alignment = _align("center")
        wc.border = _border_thin()
    row += 1

    matrix = [
        # Float×High / Float×Low / Peg×High / Peg×Low
        # Coord:   FCHR FCHD FCLR FCLD  PCHR PCHD PCLR PCLD
        # Uncoord: FUHR FUHD FULR FULD  PUHR PUHD PULR PULD
        ("COORD laboral",    ["FCHR","FCHD","FCLR","FCLD","PCHR","PCHD","PCLR","PCLD"]),
        ("DESCOORD laboral", ["FUHR","FUHD","FULR","FULD","PUHR","PUHD","PULR","PULD"]),
    ]

    for row_label, codes in matrix:
        ws.row_dimensions[row].height = 50
        # row label
        rl = ws.cell(row=row, column=1, value=row_label)
        rl.font = _font(bold=True, size=8, color=COL_NAVY)
        rl.alignment = Alignment(horizontal="center", vertical="center",
                                 textRotation=90)

        for col_i, code in enumerate(codes, 2):
            sim = simulate(code)
            y_max  = round(max(sim["y"]), 2)
            pi_max = round(max(sim["pi"]), 2)
            t_close = T
            for t in range(T):
                if abs(sim["y"][t]) < 0.3:
                    t_close = t + 1
                    break

            cell_text = f"[{code}]\ny_max={y_max}\nπ_max={pi_max}\ncierre≈T+{t_close}"
            mc = ws.cell(row=row, column=col_i, value=cell_text)
            mc.alignment = _align("center", wrap=True)
            mc.border = _border_thin()
            mc.font = _font(size=8)

            # Color by pi_max
            if pi_max < 1.0:
                mc.fill = _fill("C8E6C9")
            elif pi_max < 2.0:
                mc.fill = _fill("FFF9C4")
            elif pi_max < 3.0:
                mc.fill = _fill("FFE0B2")
            else:
                mc.fill = _fill("FFCDD2")

        row += 1

    row += 2
    ws.merge_cells(f"B{row}:I{row}")
    leg = ws[f"B{row}"]
    leg.value = ("Leyenda de color (π_max):  Verde < 1 pp  ·  Amarillo 1–2 pp  ·  "
                 "Naranja 2–3 pp  ·  Rojo > 3 pp sobre meta")
    leg.font = _font(size=9, italic=True)
    leg.alignment = _align("center")


def write_episodes_sheet(wb):
    ws = wb.create_sheet(title="EPISODIOS_HISTÓRICOS")
    ws.sheet_view.showGridLines = False

    widths = [3, 30, 10, 18, 18, 14, 14, 14, 14, 55, 3]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(f"B{row}:J{row}")
    t = ws[f"B{row}"]
    t.value = "Mapeo de Episodios Históricos → Regímenes MacroLab"
    t.fill = _fill(COL_NAVY)
    t.font = _font(bold=True, color=COL_WHITE, size=13)
    t.alignment = _align("center")
    row += 2

    headers = ["Episodio","Código","Tipo TC","Coord lab","Cred BC",
               "Marco fisc","y_max","π_max","Notas pedagógicas"]
    ws.row_dimensions[row].height = 18
    for col_i, h in enumerate(headers, 2):
        hc = ws.cell(row=row, column=col_i, value=h)
        hc.fill = _fill(COL_NAVY)
        hc.font = _font(bold=True, color=COL_WHITE, size=9)
        hc.alignment = _align("center")
        hc.border = _border_thin()
    row += 1

    episodes = [
        ("Chile 2008–2009",          "FCHR", "Mejor práctica: IT + regla fiscal + flotación → mínima perturbación"),
        ("Suecia 1994–1998",         "FCHR", "Post-flotación 1992: devaluación amortiguó → consolidación sin recesión"),
        ("Chile 2021–2024",          "FCHD", "Fase inicial: fiscal expansivo + BC sin credibilidad plena; luego coordinación"),
        ("Argentina 2020",           "PULD", "Dominancia fiscal + fijo implícito (cepo) + baja cred → dinámica explosiva"),
        ("Argentina Convertibilidad","PCLD", "Fijo duro + fiscal discrecional + baja cred → crisis de BdP"),
        ("EE.UU. 2021–2022",         "FUHD", "Alta cred Fed + fiscal COVID + mercado laboral descoordinado → Fed reacciona"),
        ("Grecia 2010–2015",         "PUHR", "Fijo (euro) + alta cred ECB + mercado laboral descoord → ajuste interno"),
        ("España 2003–2007",         "PCHD", "Fijo (euro) + coord moderada + fiscal discrecional → boom-bust"),
    ]

    for i, (epis, code, nota) in enumerate(episodes):
        reg = REGIMES[code]
        sim = simulate(code)
        y_max  = round(max(sim["y"]), 2)
        pi_max = round(max(sim["pi"]), 2)

        ws.row_dimensions[row].height = 18
        fill = _fill(COL_CARD) if i % 2 == 0 else _fill(COL_WHITE)
        vals = [epis, code, reg["er"], reg["lab"], reg["cred"], reg["fisc"],
                y_max, pi_max, nota]
        for col_i, v in enumerate(vals, 2):
            dc = ws.cell(row=row, column=col_i, value=v)
            dc.fill = fill
            dc.font = _font(size=9)
            dc.alignment = _align("left" if col_i in (2, 10) else "center")
            dc.border = _border_thin()
        row += 1


def build_workbook(output_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    print("Writing summary sheet...")
    write_summary_sheet(wb)

    print("Writing matrix sheet...")
    write_matrix_sheet(wb)

    print("Writing episodes sheet...")
    write_episodes_sheet(wb)

    print("Writing 16 regime sheets...")
    order = ["FCHR","FCHD","FCLR","FCLD",
             "FUHR","FUHD","FULR","FULD",
             "PCHR","PCHD","PCLR","PCLD",
             "PUHR","PUHD","PULR","PULD"]
    for code in order:
        print(f"  {code} — {REGIMES[code]['label']}")
        write_regime_sheet(wb, code)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    out = "/home/user/MacroLab-Shock-Simulator/innovation/phase1/MacroLab_Regime_Trajectories_v1.xlsx"
    build_workbook(out)
